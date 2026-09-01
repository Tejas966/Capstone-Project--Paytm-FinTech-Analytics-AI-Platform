"""
Part 1A — Excel Merchant Workbook Generator
Paytm Payments Analytics

Generates merchant_workbook.xlsx with:
  Sheet 1 (Merchants)     : raw merchants data (reference table)
  Sheet 2 (Transactions)  : ledger with VLOOKUP columns + nested IF/AND classification
                            Rule: "High-Value Merchant Day" when merchant's DAILY
                            transaction total > INR 5,000 AND region != "East"
                            Uses SUMPRODUCT to compute the daily total across all
                            transactions for the same merchant on the same day.
  Sheet 3 (HLOOKUP_Demo)  : horizontally-laid MDR fee-rate table + HLOOKUP demo
  Sheet 4 (Pivot_Summary) : real Excel PivotTable (merchant_id x status) via xlwings
  Sheet 5 (Unique_Days)   : qualifying merchant-day totals (not averages) + classification

Run from inside payments_fraud_analytics/:
    cd payments_fraud_analytics && python generate_workbook.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import os

# Styling constants
BLUE_DARK   = "1D4ED8"
BLUE_MID    = "3B82F6"
AMBER       = "D97706"
WHITE       = "FFFFFF"
GRAY_HEADER = "1E293B"
LIGHT_BLUE  = "DBEAFE"
LIGHT_GREEN = "DCFCE7"
LIGHT_GRAY  = "F8FAFC"
LIGHT_AMBER = "FEF3C7"

thin_border = Border(
    left   = Side(style="thin", color="CBD5E1"),
    right  = Side(style="thin", color="CBD5E1"),
    top    = Side(style="thin", color="CBD5E1"),
    bottom = Side(style="thin", color="CBD5E1"),
)

def hdr_style(cell, bg=GRAY_HEADER, fg=WHITE, bold=True, size=10):
    cell.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border

def data_style(cell, bg=WHITE, fg="1E293B", bold=False):
    cell.font      = Font(bold=bold, color=fg, size=9, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = thin_border

def num_style(cell, bg=WHITE, fg="1E293B"):
    cell.font      = Font(color=fg, size=9, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border    = thin_border

def section_title(ws, row, col, text, span=1, bg=BLUE_DARK):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(bold=True, color=WHITE, size=11, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)

# Load data
merchants_df = pd.read_csv("merchants.csv")
ledger_df    = pd.read_csv("ledger.csv", parse_dates=["transaction_time"])

wb = openpyxl.Workbook()
wb.remove(wb.active)

# =========================================================================
# SHEET 1 — Merchants (reference table for VLOOKUP)
# =========================================================================
ws_m = wb.create_sheet("Merchants")

m_headers = ["merchant_id", "merchant_name", "category", "region"]
for col, h in enumerate(m_headers, 1):
    hdr_style(ws_m.cell(row=1, column=col, value=h), bg=BLUE_DARK)

for r_idx, row in enumerate(merchants_df.itertuples(index=False), start=2):
    bg = LIGHT_BLUE if r_idx % 2 == 0 else WHITE
    for col, val in enumerate([row.merchant_id, row.merchant_name,
                                row.category, row.region], 1):
        cell = ws_m.cell(row=r_idx, column=col, value=val)
        if col == 1:
            num_style(cell, bg=bg)
        else:
            data_style(cell, bg=bg)

ws_m.column_dimensions["A"].width = 14
ws_m.column_dimensions["B"].width = 18
ws_m.column_dimensions["C"].width = 18
ws_m.column_dimensions["D"].width = 12
ws_m.freeze_panes = "A2"

# =========================================================================
# SHEET 2 — Transactions (VLOOKUP + nested IF with SUMPRODUCT daily total)
# =========================================================================
ws_t = wb.create_sheet("Transactions")

t_headers = [
    "transaction_id", "user_id", "merchant_id", "transaction_time",
    "amount_inr", "payment_method", "status", "risk_score",
    "merchant_name (VLOOKUP)", "category (VLOOKUP)", "region (VLOOKUP)",
    "classification (nested IF)"
]
for col, h in enumerate(t_headers, 1):
    bg = GRAY_HEADER if col <= 8 else BLUE_MID if col <= 11 else AMBER
    hdr_style(ws_t.cell(row=1, column=col, value=h), bg=bg)

MERCH_RANGE = "Merchants!$A$2:$D$41"
n_data      = len(ledger_df)
last_row    = n_data + 1   # header row 1, data rows 2..last_row

for r_idx, row in enumerate(ledger_df.itertuples(index=False), start=2):
    bg  = LIGHT_BLUE if r_idx % 2 == 0 else WHITE
    bg2 = "EFF6FF"   if r_idx % 2 == 0 else "F0F9FF"
    bg3 = LIGHT_AMBER if r_idx % 2 == 0 else "FFFBEB"

    vals = [row.transaction_id, row.user_id, row.merchant_id,
            str(row.transaction_time), row.amount_inr,
            row.payment_method, row.status, row.risk_score]

    for col, val in enumerate(vals, 1):
        cell = ws_t.cell(row=r_idx, column=col, value=val)
        if col in (1, 6, 7):
            data_style(cell, bg=bg)
        else:
            num_style(cell, bg=bg)

    # VLOOKUP formulas (columns 9, 10, 11)
    for out_col, vlookup_col_idx in [(9, 2), (10, 3), (11, 4)]:
        formula = (f'=IFERROR(VLOOKUP($C{r_idx},'
                   f'{MERCH_RANGE},{vlookup_col_idx},FALSE),"Merchant not found")')
        cell = ws_t.cell(row=r_idx, column=out_col, value=formula)
        data_style(cell, bg=bg2)

    # Nested IF/AND classification (column 12)
    # Rule: "High-Value Merchant Day" when:
    #   merchant's daily transaction total > INR 5,000
    #   AND region (from VLOOKUP) is not "East"
    #
    # SUMPRODUCT computes the sum of amount_inr for all transactions
    # with the same merchant_id on the same calendar day:
    #   SUMPRODUCT(($C$2:$C${last}=$C{row})
    #             *(LEFT($D$2:$D${last},10)=LEFT($D{row},10))
    #             *$E$2:$E${last})
    sumproduct = (
        f'SUMPRODUCT(($C$2:$C${last_row}=$C{r_idx})'
        f'*(LEFT($D$2:$D${last_row},10)=LEFT($D{r_idx},10))'
        f'*$E$2:$E${last_row})'
    )
    formula_if = (
        f'=IF(AND({sumproduct}>5000,'
        f'$K{r_idx}<>"East"),'
        f'"High-Value Merchant Day","Standard")'
    )
    cell = ws_t.cell(row=r_idx, column=12, value=formula_if)
    data_style(cell, bg=bg3)

col_widths = [15, 9, 13, 22, 12, 16, 12, 12, 22, 18, 14, 26]
for col, w in enumerate(col_widths, 1):
    ws_t.column_dimensions[get_column_letter(col)].width = w
ws_t.freeze_panes = "A2"
ws_t.row_dimensions[1].height = 28

note_cell = ws_t.cell(row=1, column=14, value=(
    "VLOOKUP uses fixed range Merchants!$A$2:$D$41 with $ absolute references. "
    "IFERROR shows 'Merchant not found' for unmatched merchant_id. "
    "Nested IF rule: merchant's daily total (via SUMPRODUCT) > INR 5,000 "
    "AND region != 'East' -> 'High-Value Merchant Day'."
))
note_cell.font = Font(italic=True, color="64748B", size=8)

# =========================================================================
# SHEET 3 — HLOOKUP Demo (horizontally-laid MDR fee-rate table)
# =========================================================================
ws_h = wb.create_sheet("HLOOKUP_Demo")

section_title(ws_h, 1, 1, "MDR Fee-Rate Reference Table (Horizontal Layout)", span=5, bg=BLUE_DARK)

methods = ["UPI", "Wallet", "Card", "Netbanking"]
fee_pct = [0.25,   0.50,    1.80,    0.90]

ws_h.cell(row=2, column=1, value="Payment Method")
for col, m in enumerate(methods, 2):
    hdr_style(ws_h.cell(row=2, column=col, value=m), bg=BLUE_MID)
hdr_style(ws_h.cell(row=2, column=1, value="Payment Method"), bg=GRAY_HEADER)

ws_h.cell(row=3, column=1, value="MDR Fee %")
hdr_style(ws_h.cell(row=3, column=1, value="MDR Fee %"), bg=GRAY_HEADER)
for col, fee in enumerate(fee_pct, 2):
    cell = ws_h.cell(row=3, column=col, value=fee)
    cell.number_format = "0.00%"
    data_style(cell, bg=LIGHT_BLUE)

ws_h.cell(row=4, column=1, value="")

section_title(ws_h, 5, 1, "HLOOKUP Demonstration", span=5, bg=AMBER)

ws_h.cell(row=6, column=1, value="Lookup Method")
ws_h.cell(row=7, column=1, value="HLOOKUP Formula")
ws_h.cell(row=8, column=1, value="Result")

for cell_ref in [ws_h.cell(row=6, column=1), ws_h.cell(row=7, column=1),
                 ws_h.cell(row=8, column=1)]:
    hdr_style(cell_ref, bg=GRAY_HEADER)

lookup_examples = ["UPI", "Wallet", "Card", "Netbanking"]
for col, method in enumerate(lookup_examples, 2):
    cell_in = ws_h.cell(row=6, column=col, value=method)
    data_style(cell_in, bg=LIGHT_AMBER)
    cell_in.font = Font(bold=True, color="92400E", size=10)

    col_letter = get_column_letter(col)
    formula = f'=HLOOKUP({col_letter}6,$B$2:$E$3,2,FALSE)'
    cell_fml = ws_h.cell(row=7, column=col, value=formula)
    data_style(cell_fml, bg="FFF7ED")
    cell_fml.font = Font(italic=True, color="92400E", size=9)

    hlookup_result = f'=HLOOKUP({col_letter}6,$B$2:$E$3,2,FALSE)'
    cell_res = ws_h.cell(row=8, column=col, value=hlookup_result)
    cell_res.number_format = "0.00%"
    data_style(cell_res, bg=LIGHT_GREEN)
    cell_res.font = Font(bold=True, color="166534", size=10)

note = ws_h.cell(row=10, column=1,
    value=("Design decision: MDR rates are illustrative industry approximations. "
           "UPI 0.25% (RBI-capped for small merchants), Wallet 0.50%, "
           "Card 1.80% (average interchange), Netbanking 0.90% (bank surcharge). "
           "HLOOKUP uses row_index=2 to return the fee % from the second row of the table."))
note.font = Font(italic=True, color="64748B", size=8)
ws_h.merge_cells("A10:E10")
note.alignment = Alignment(wrap_text=True)
ws_h.row_dimensions[10].height = 48

for col_letter, w in [("A", 22), ("B", 14), ("C", 14), ("D", 14), ("E", 14)]:
    ws_h.column_dimensions[col_letter].width = w

# =========================================================================
# SHEET 5 — Unique Days (qualifying merchant-day totals + classification)
# =========================================================================
ws_u = wb.create_sheet("Unique_Days")

section_title(ws_u, 1, 1,
    "Qualifying Merchant-Day Totals & Classification (Top 10 Merchants)",
    span=10, bg=BLUE_DARK)

ledger_df["txn_date"] = ledger_df["transaction_time"].dt.date

# Compute per-merchant-per-day totals (the actual daily totals, not averages)
merchant_day_totals = ledger_df.groupby(["merchant_id", "txn_date"]).agg(
    daily_total_inr=("amount_inr", "sum"),
    daily_txn_count=("transaction_id", "count")
).reset_index()

# Per-merchant summary with qualifying day counts
merchant_unique = merchant_day_totals.groupby("merchant_id").agg(
    unique_days                =("txn_date",        "nunique"),
    total_gmv_inr              =("daily_total_inr", "sum"),
    qualifying_high_value_days =("daily_total_inr", lambda x: int((x > 5000).sum())),
    max_daily_total            =("daily_total_inr", "max"),
    total_txn_count            =("daily_txn_count", "sum"),
).reset_index()

merchant_unique = merchant_unique.merge(
    merchants_df[["merchant_id", "merchant_name", "category", "region"]],
    on="merchant_id"
)

# Classification: has any qualifying high-value days AND region != East
merchant_unique["classification"] = merchant_unique.apply(
    lambda r: "High-Value Merchant Day"
    if r["qualifying_high_value_days"] > 0 and r["region"] != "East"
    else "Standard",
    axis=1
)

top10_unique = merchant_unique.nlargest(10, "total_txn_count")

u_headers = [
    "merchant_id", "merchant_name", "category", "region",
    "total_txn_count", "unique_days", "total_gmv_inr (INR)",
    "qualifying_days (daily > 5000)", "max_daily_total (INR)",
    "Classification (IF rule)"
]
for col, h in enumerate(u_headers, 1):
    bg = GRAY_HEADER if col < 5 else BLUE_MID if col <= 7 else AMBER
    hdr_style(ws_u.cell(row=2, column=col, value=h), bg=bg)

for r_idx, row in enumerate(top10_unique.itertuples(index=False), start=3):
    bg     = LIGHT_BLUE if r_idx % 2 != 0 else WHITE
    bg_cls = LIGHT_GREEN if row.classification == "High-Value Merchant Day" \
             else LIGHT_GRAY
    vals = [
        row.merchant_id, row.merchant_name, row.category, row.region,
        row.total_txn_count, row.unique_days, row.total_gmv_inr,
        row.qualifying_high_value_days, row.max_daily_total,
        row.classification
    ]
    for col, val in enumerate(vals, 1):
        cell = ws_u.cell(row=r_idx, column=col, value=val)
        if col == 10:
            data_style(cell, bg=bg_cls,
                       fg="166534" if "High" in str(val) else "1E293B",
                       bold="High" in str(val))
        elif col in (1, 2, 3, 4):
            data_style(cell, bg=bg)
        else:
            num_style(cell, bg=bg)

rule_note = ws_u.cell(row=14, column=1,
    value=("Nested IF/AND Classification Rule: 'High-Value Merchant Day' when "
           "merchant's DAILY transaction total (sum of all txns for that merchant "
           "on that day) > INR 5,000 AND region != 'East'. "
           "'qualifying_days' counts how many calendar days exceeded this threshold. "
           "'max_daily_total' shows the highest single-day total for the merchant."))
rule_note.font = Font(italic=True, color="64748B", size=8)
ws_u.merge_cells("A14:J14")
rule_note.alignment = Alignment(wrap_text=True)
ws_u.row_dimensions[14].height = 42

col_widths_u = [13, 18, 16, 12, 16, 14, 18, 28, 22, 28]
for col, w in enumerate(col_widths_u, 1):
    ws_u.column_dimensions[get_column_letter(col)].width = w
ws_u.freeze_panes = "A3"

# =========================================================================
# Save with openpyxl (sheets 1, 2, 3, 5 — no Pivot_Summary yet)
# =========================================================================
OUTFILE = "merchant_workbook.xlsx"
wb.save(OUTFILE)
print(f"Phase 1 complete: {OUTFILE} saved (Merchants, Transactions, HLOOKUP_Demo, Unique_Days)")

# =========================================================================
# SHEET 4 — Real Excel PivotTable via xlwings COM automation
# =========================================================================
print("Phase 2: Creating real Excel PivotTable via xlwings...")

try:
    import xlwings as xw

    app = xw.App(visible=False, add_book=False)
    try:
        wb_xw = app.books.open(os.path.abspath(OUTFILE))

        # Create Pivot_Summary sheet after HLOOKUP_Demo
        pivot_sheet = wb_xw.sheets.add("Pivot_Summary",
                                        after=wb_xw.sheets["HLOOKUP_Demo"])

        # Title cell
        pivot_sheet.range("A1").value = "Pivot Table: Total Amount & Count by Merchant x Status"
        pivot_sheet.range("A1").font.bold = True
        pivot_sheet.range("A1").font.size = 12

        # Source data: Transactions sheet, columns A:H (raw data only)
        src_sheet = wb_xw.sheets["Transactions"]
        src_last_row = src_sheet.range("A1").end("down").row
        source_address = f"Transactions!$A$1:$H${src_last_row}"

        # Create PivotCache (xlDatabase = 1)
        pivot_cache = wb_xw.api.PivotCaches().Create(
            SourceType=1,
            SourceData=source_address
        )

        # Create PivotTable at cell A3 on the Pivot_Summary sheet
        pivot_table = pivot_cache.CreatePivotTable(
            TableDestination=pivot_sheet.api.Cells(3, 1),
            TableName="MerchantStatusPivot"
        )

        # Row field: merchant_id
        pf_merchant = pivot_table.PivotFields("merchant_id")
        pf_merchant.Orientation = 1   # xlRowField
        pf_merchant.Position = 1

        # Column field: status
        pf_status = pivot_table.PivotFields("status")
        pf_status.Orientation = 2     # xlColumnField
        pf_status.Position = 1

        # Data field 1: Sum of amount_inr
        pivot_table.AddDataField(
            pivot_table.PivotFields("amount_inr"),
            "Total Amount (INR)",
            -4157   # xlSum
        )

        # Data field 2: Count of transaction_id
        pivot_table.AddDataField(
            pivot_table.PivotFields("transaction_id"),
            "Txn Count",
            -4112   # xlCount
        )

        wb_xw.save()
        wb_xw.close()
        print("Real Excel PivotTable created successfully on Pivot_Summary sheet.")

    finally:
        app.quit()

except ImportError:
    print("WARNING: xlwings not installed. Pivot_Summary sheet NOT created.")
    print("Install xlwings (pip install xlwings) and re-run to add the PivotTable.")
except Exception as e:
    print(f"WARNING: Could not create PivotTable via Excel COM: {e}")
    print("The workbook was saved without the PivotTable sheet.")
    print("You can manually add one in Excel using the Transactions data.")

print(f"\n{OUTFILE} saved successfully.")
print("Sheets: Merchants | Transactions | HLOOKUP_Demo | Pivot_Summary | Unique_Days")
